##########

from Tkinter import *
import ttk
from tkMessageBox import *
import datetime, time
import os
from PIL import Image, ImageTk
import PIL
from xlrd import open_workbook
import xlrd
import math
from tkFileDialog import askopenfilename,askdirectory
#import sys
#sys.path.insert(0, "/home/user/PycharmProjects/gui/mpldatacursor")
from mpldatacursor import datacursor
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import axes3d
from openpyxl import *
import cv2
import pyexiv2

#######

class REG:

    def __init__(self, master):
        self.master = master
        self.master.title("  Trajectory-Frame Extraction & Geotagging ")
        self.master.config(bg="lavender")
        self.master.geometry("800x450+35+80")

        self.valuesdata = StringVar()
        self.videodata = StringVar()
        self.intiallat = StringVar()
        self.endlat = StringVar()
        #self.excelfilename = StringVar()
        self.outputfoldername = StringVar()
        self.plottype = StringVar()
        self.numberofframe=StringVar()
        self.coltime=StringVar()
        self.collat=StringVar()
        self.collong=StringVar()
        self.colalt=StringVar()




        self.can = Canvas(master, width=1000, height=100, bg='white')

        self.can.place(x=1, y=1)

        self.label =Label(master, text="Processing Box",bg="black",font=('verdana', 10, 'bold'),fg="white",borderwidth=2,relief="groove",width="70",height="3")
        self.label.place(x=150,y=390)

        self.l = Label(master, text="Frame Extraction Module", bg='white', fg='MidnightBlue',
                       font=('Helvetica', 15, 'bold italic'))
        self.l.place(x=280, y=25)
        self.labelline = Label(master,
                               text="-----------------------------------------------------------------------------------------------------------------------------------------------------------------------",
                               bg='Lavender', fg="black", font=('verdana', 12, 'bold'))
        self.labelline.place(x=1, y=197)

        self.ltime = Label(master, text="Time", bg='Lavender', fg='red')
        self.ltime.place(x=530, y=105)
        self.cmbtime = ttk.Combobox(master, textvariable=self.coltime)
        self.cmbtime["values"] = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
        self.cmbtime.current(0)
        self.cmbtime.place(x=530, y=122, width=50)

        self.llat = Label(master, text="Lat", bg='Lavender', fg='red')
        self.llat.place(x=590, y=105)
        self.cmblat = ttk.Combobox(master, textvariable=self.collat)
        self.cmblat["values"] = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
        self.cmblat.current(1)
        self.cmblat.place(x=590, y=122, width=50)

        self.llong = Label(master, text="Long", bg='Lavender', fg='red')
        self.llong.place(x=650, y=105)
        self.cmblong = ttk.Combobox(master, textvariable=self.collong)
        self.cmblong["values"] = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
        self.cmblong.current(2)
        self.cmblong.place(x=650, y=122, width=50)

        self.lalt = Label(master, text="Altitude", bg='Lavender', fg='red')
        self.lalt.place(x=710, y=105)
        self.cmbalt = ttk.Combobox(master, textvariable=self.colalt)
        self.cmbalt["values"] = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
        self.cmbalt.current(3)
        self.cmbalt.place(x=710, y=122, width=50)

        self.e1 = Label(master, bg='white', textvariable=self.valuesdata, )
        self.e1.place(x=150, y=120, width=350)
        self.l1 = Label(master, text="Intial Latitude /Time(x):", bg='Lavender', font=('verdana', 10, 'bold'))
        self.l1.place(x=10, y=250)
        self.e2 = Entry(master, textvariable=self.intiallat, validate="key", )
        self.e2['validatecommand'] = (self.e2.register(self.testintiallatVal), '%P', '%i', '%d')
        self.e2.place(x=200, y=250, width=250)

        self.l2 = Label(master, text="End Latitude / Time(x):", bg='Lavender', font=('verdana', 10, 'bold'))
        self.l2.place(x=10, y=300)
        self.e3 = Entry(master, textvariable=self.endlat, validate="key")
        self.e3['validatecommand'] = (self.e3.register(self.testendlatVal), '%P', '%i', '%d')
        self.e3.place(x=200, y=300, width=250)
        '''
        self.l3 = Label(master, text="Excel file name:", bg='Lavender', font=('verdana', 10, 'bold'))
        self.l3.place(x=10, y=350)
        self.e4 = Entry(master, textvariable=self.excelfilename, )
        self.e4.place(x=200, y=350, width=250)
        '''
        self.e6 = Label(master, bg='white',textvariable=self.outputfoldername, )
        self.e6.place(x=150, y=350, width=350)

        self.e7 = Label(master, bg='white', textvariable=self.videodata, )
        self.e7.place(x=150, y=220, width=350)

        self.l4 = Label(master, text="Plot Mode:", bg='Lavender', font=('verdana', 10, 'bold'))
        self.l4.place(x=10, y=160)
        self.cmb1 = ttk.Combobox(master, text=self.plottype, state='readonly')
        self.cmb1["values"] = ["Time & Altitude(2D)", "Latitude,Longitude & Altitude(3d)"]
        self.cmb1.current(0)
        self.cmb1.place(x=150, y=160, width=250)
        self.style = ttk.Style()
        self.style.map('TCombobox', fieldbackground=[('readonly', 'white')])
        self.style.map('TCombobox', selectbackground=[('readonly', 'white')])
        self.style.map('TCombobox', selectforeground=[('readonly', 'red')])

        ############
        self.l5 = Label(master, text="Frame/Second", bg='Lavender', font=('verdana', 10, 'bold'))
        self.l5.place(x=500, y=220)
        self.cmb2 = ttk.Combobox(master, text=self.numberofframe)
        self.cmb2["values"] = ["All", "1"]
        self.cmb2.current(1)
        self.cmb2.place(x=630, y=220, width=100)
        self.style2 = ttk.Style()
        self.style2.map('TCombobox', fieldbackground=[('pressed','white')])
        self.style2.map('TCombobox', selectbackground=[('pressed','white')])
        self.style2.map('TCombobox', selectforeground=[('pressed','red')])


        self.b1 = Button(master, text='Import Excel File', cursor="plus", bg='brown', fg='white', command=self.callback)
        self.b1.place(x=5, y=120, width=110)

        self.b2 = Button(master, text=" Plot  ", command=self.fnplot, bg='Midnight Blue', fg='white',
                         font=('verdana', 10, 'bold'), bd=1, width=15, height=2)
        self.b2.place(x=620, y=165)

        self.b3 = Button(master, text=" Extract ", command=self.fnsubplot, bg='Midnight Blue', fg='white',
                         font=('verdana', 10, 'bold'), bd=1, width=15, height=2)
        self.b3.place(x=620, y=270)

        self.b4 = Button(master, text="Cancel", command=self.fncancel, bg='Midnight Blue', fg='white',
                         font=('verdana', 10, 'bold'), bd=1, width=15, height=2)
        self.b4.place(x=620, y=340)

        self.b5 = Button(text="Select output folder", cursor="plus", bg='brown', fg='white', command=self.browse_button)
        self.b5.place(x=5, y=350)

        self.b6 = Button(master, text='Import Video', cursor="plus", bg='brown', fg='white', command=self.callbackvideo)
        self.b6.place(x=5, y=220, width=110)

        self.pi = PIL.Image.open('//home//user//PycharmProjects//gui//isrologo.jpg')
        self.im = ImageTk.PhotoImage(self.pi)
        self.picl = Label(master, image=self.im, height=90, width=96)
        self.picl.place(x=30, y=5)

        self.pi2 = PIL.Image.open('//home//user//PycharmProjects//gui//IIRS_logo.jpg')
        self.im2 = ImageTk.PhotoImage(self.pi2)
        self.picl2 = Label(master, image=self.im2, height=90, width=180)
        self.picl2.place(x=600, y=1)

    def callback(self):
        name = askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.valuesdata.set(name)

    def callbackvideo(self):
        name1 = askopenfilename(filetypes=[("Mp4 files", "*.mp4")])
        self.videodata.set(name1)

    def testintiallatVal(self, inStr, i, acttyp):
        ind = int(i)
        if acttyp == '1':  # insert
            if inStr[ind] != '.':

                if not inStr[ind].isdigit():
                    return False
        return True

    def testendlatVal(self, inStr1, j, acttyp1):
        ind1 = int(j)
        if acttyp1 == '1':  # insert
            if inStr1[ind1] != '.':

                if not inStr1[ind1].isdigit():
                    return False
        return True

    def browse_button(self):
        # Allow user to select a directory and store it in global var
        # called folder_path
        global folder_path
        filename = askdirectory()
        self.outputfoldername.set(filename)

    def to_deg(self,value, loc):
        if value < 0:
            loc_value = loc[0]
        elif value > 0:
            loc_value = loc[1]
        else:
            loc_value = ""
        abs_value = abs(value)
        deg = int(abs_value)
        t1 = (abs_value - deg) * 60
        min = int(t1)
        sec = round((t1 - min) * 60, 5)
        return (deg, min, sec, loc_value)

    def set_gps_location(self,file_name, lat, lng,alt):
        """Adds GPS position as EXIF metadata

           Keyword arguments:
           file_name -- image file
           lat -- latitude (as float)
           lng -- longitude (as float)

           """
        lat_deg = self.to_deg(lat, ["S", "N"])
        lng_deg = self.to_deg(lng, ["W", "E"])



        # convert decimal coordinates into degrees, munutes and seconds
        exiv_lat = (pyexiv2.Rational(lat_deg[0] * 60 + lat_deg[1], 60), pyexiv2.Rational(lat_deg[2] * 100, 6000),
                    pyexiv2.Rational(0, 1))
        exiv_lng = (pyexiv2.Rational(lng_deg[0] * 60 + lng_deg[1], 60), pyexiv2.Rational(lng_deg[2] * 100, 6000),
                    pyexiv2.Rational(0, 1))
        metadata = pyexiv2.ImageMetadata(file_name)

        metadata.read()
        ##    exif_keys = metadata.exif_keys

        metadata["Exif.GPSInfo.GPSLatitude"] = exiv_lat
        metadata["Exif.GPSInfo.GPSLatitudeRef"] = lat_deg[3]
        metadata["Exif.GPSInfo.GPSLongitude"] = exiv_lng
        metadata["Exif.GPSInfo.GPSLongitudeRef"] = lng_deg[3]
        metadata["Exif.GPSInfo.GPSAltitude"]=pyexiv2.Rational(alt,1)
        metadata["Exif.Image.GPSTag"] = 654
        metadata["Exif.GPSInfo.GPSMapDatum"] = "WGS-84"
        metadata["Exif.GPSInfo.GPSVersionID"] = '2 0 0 0'
        metadata.write()

    def fnplot(self):


        if (self.valuesdata.get() == ""):
            showerror("Error", "Excel File Must be Import")
            return

        self.a = self.valuesdata.get()
        book = xlrd.open_workbook(self.a)
        self.label.config(text="Reading.."+str(self.valuesdata.get()),fg="red")
        self.label.update_idletasks()
        sheet = book.sheet_by_index(0)  # If your data is on sheet 1
        plotmode = self.plottype.get()
        if (plotmode == 'Latitude,Longitude & Altitude(3d)'):

            columnA = []  # this is for lattitude
            columnB = []  # this is for longitude
            columnC = []  # this is for Altitude

            columnno1 = int(self.collat.get())
            columnno2 = int(self.collong.get())
            columnno3 = int(self.colalt.get())
            for row in range(1, sheet.nrows):  # start from 1, to leave out row 0

                columnA.append(sheet.cell_value(row, columnno1))  # extract from Latitude col
                columnB.append(sheet.cell_value(row, columnno2))  # extract from Longitude col
                columnC.append(sheet.cell_value(row, columnno3))  # extract from Altitude col

            fig = plt.figure()
            ax = fig.gca(projection='3d')
            surf = ax.plot(columnA, columnB, columnC)
            datacursor(surf)
            ax.set_xlabel('Lat')
            ax.set_ylabel('Long')
            ax.set_zlabel('Altitude')
            plt.show()
            self.label.config(text="Trajectory Plotted")
            self.label.update_idletasks()
        else:
            columnA = []  # time
            columnB = []  # Altitude

            columnno0 = int(self.coltime.get())
            columnno3 = int(self.colalt.get())

            for row in range(1, sheet.nrows):  # start from 1, to leave out row 0

                columnA.append(sheet.cell_value(row, columnno0))  # extract from Time col
                columnB.append(sheet.cell_value(row, columnno3))  # extract from Altitude col

            surf = plt.plot(columnA, columnB)
            datacursor(surf)
            plt.xlabel('Time')

            plt.ylabel('Altitude')
            plt.show()
            self.label.config(text=" Trajectory Plotted")
            self.label.update_idletasks()

    def fncancel(self):
        self.valuesdata.set("")
        self.videodata.set("")
        self.intiallat.set("")
        self.endlat.set("")
        self.outputfoldername.set("")
        self.numberofframe.set("1")
        self.label.config(text="Processing Box")
        self.label.update_idletasks()

    def fnsubplot(self):
        columnno0 = int(self.coltime.get())
        columnno1 = int(self.collat.get())
        columnno2 = int(self.collong.get())
        columnno3 = int(self.colalt.get())



        if (self.valuesdata.get() == "" or self.videodata.get() == "" or self.intiallat.get() == "" or self.endlat.get() == "" or self.outputfoldername.get() == ""):
            showerror("Error", "All Entries Must be Filled")
            return


        # set of list to store data
        column0 = []
        column1 = []
        column2 = []
        column3 = []
        column4 = []
        columnimagename = []

        # reading excel file

        f = self.valuesdata.get()

        global sheet, min, max




        self.label.config(text="Processing sheet."+f,fg="red")
        self.label.update_idletasks()

        # looping through all rows and first column

        a = float(self.intiallat.get())
        b = float(self.endlat.get())

        if a < b:
             min = a
             max = b


        else:
             min = b
             max = a


        # Finding the data between the given latitude

        global cell_row1, cell_row11, cell_row2, cell_row22
        global validateendtime
        global validateintialtime
        validateintialtime = 0
        validateendtime = 0
        plotmode = self.plottype.get()




        if (plotmode == 'Latitude,Longitude & Altitude(3d)'):



            self.label.config(text="Extracting data from worksheet & Video file...")
            self.label.update_idletasks()
            '''
            for row in sheet.iter_rows(
                    'B{}:B{}'.format(2, sheet.max_row)):  # here B defines the column you want to search
                for col in sheet.iter_cols(min_col=1,
                                           max_col=1):  # min column and maaximum column diffrence define the repetion of cell value
                    for cell in row:
                        # printing all cell values from all rows and first column
                        cv = cell.value

                        if cv >= min and cv <= max:
                            column1.append(cell.value)
            '''
            book = xlrd.open_workbook(f)
            sheet = book.sheet_by_index(0)
            for row in range(1, sheet.nrows):

                cv = sheet.cell_value(row, columnno1)
                if (cv >= min and cv <= max):
                    column1.append(sheet.cell_value(row, columnno1))

            ######################################################################################################################

            column1 = sorted(column1, key=float)
            intial_Lat = column1[0]
            l = len(column1)
            end_Lat = column1[l - 1]




            # finding the row number of first latitude and last latitude
            '''
            for row in sheet.iter_rows(
                    'B{}:B{}'.format(2, sheet.max_row)):  # here B defines the column you want to search
                for col in sheet.iter_cols(min_col=1,
                                           max_col=1):  # min column and maaximum column diffrence define the repetion of cell value
                    for cell in row:
                        cv = cell.value
                        if cv == intial_Lat:
                            cell_row1 = cell.row
                            cell_row11 = cell_row1 - 1

                        if cv == end_Lat:
                            cell_row2 = cell.row
                            cell_row22 = cell_row2 - 1
            '''
            for row in range(1, sheet.nrows):
                cv = (sheet.cell_value(row, columnno1))
                if (cv == intial_Lat):
                    cell_row1 = row
                    cell_row11 = cell_row1 - 1
                    validateintialtime = 1

                if (cv == end_Lat):
                    cell_row2 = row
                    cell_row22 = cell_row2 - 1
                    validateendtime = 1


            if (validateintialtime != 1):
                showerror("Error", "Intial time is out of range")
                self.e2.focus_set()
                return

            if (validateendtime != 1):
                showerror("Error", "End time is out of range")
                self.e3.focus_set()
                return

            # Reading imported excel file and get data from excel sheet to list on the basis of list value

            book = xlrd.open_workbook(f)

            sheet = book.sheet_by_index(0)

            for row in range(cell_row22, cell_row1):  # start from 1, to leave out row 0

                column2.append(sheet.cell_value(row, columnno1))  # extract from Latitude col
                column3.append(sheet.cell_value(row, columnno2))  # extract from Longitude col
                column4.append(sheet.cell_value(row, columnno3))  # extract from Altitude col
            '''
            wb = load_workbook(f)
            global sheet1
            for sheet1 in wb:
                print("\n")

            for row in sheet1.iter_rows(
                    'A{}:A{}'.format(cell_row2, cell_row1)):  # here A defines the column you want to search
                for col in sheet1.iter_cols(min_col=1,
                                            max_col=1):  # min column and maaximum column diffrence define the repetion of cell value
                    for cell in row:
                        # printing all cell values from all rows and first column
                        cv = cell.value
                        column0.append(cell.value)
            '''
            book = xlrd.open_workbook(f)
            sheet = book.sheet_by_index(0)
            for row in range(cell_row22, cell_row1):
                column0.append(str(datetime.timedelta(sheet.cell_value(row, columnno0))))



        else:


            self.label.config(text="Extracting Data from worksheet & Video File",fg="red")
            self.label.update_idletasks()

            intialtime = str(datetime.timedelta(min))
            intial, el = intialtime.split(".")


            endtime = str(datetime.timedelta(max))
            end, la = endtime.split(".")


            book = xlrd.open_workbook(f)
            sheet = book.sheet_by_index(0)
            '''
            for row in sheet.iter_rows(
                    'A{}:A{}'.format(2, sheet.max_row)):  # here A defines the column you want to search
                for col in sheet.iter_cols(min_col=1,
                                           max_col=1):  # min column and maaximum column diffrence define the repetion of cell value
                    for cell in row:
                        cv = cell.value

                        if str(cv) == str(intial):
                            cell_row1 = cell.row
                            cell_row11 = cell_row1 - 1
                            validateintialtime = 1

                        if str(cv) == str(end):
                            cell_row2 = cell.row
                            cell_row22 = cell_row2 - 1
                            validateendtime = 1
             '''
            for row in range(1, sheet.nrows):
                cv = str(datetime.timedelta(sheet.cell_value(row, columnno0)))
                if str(cv) == str(intial):
                    cell_row1 = row
                    cell_row11 = cell_row1 - 1
                    validateintialtime = 1

                if str(cv) == str(end):
                    cell_row2 = row
                    cell_row22 = cell_row2 - 1
                    validateendtime = 1

            if (validateintialtime != 1):
                showerror("Error", "Intial time is out of range")
                self.e2.focus_set()
                return

            if (validateendtime != 1):
                showerror("Error", "End time is out of range")
                self.e3.focus_set()
                return

            # Reading imported excel file and get data from excel sheet to list on the basis of list value

            book = xlrd.open_workbook(f)

            sheet = book.sheet_by_index(0)

            for row in range(cell_row11, cell_row2):  # start from 1, to leave out row 0

                column2.append(sheet.cell_value(row, columnno1))  # extract from Latitude col
                column3.append(sheet.cell_value(row, columnno2))  # extract from Longitude col
                column4.append(sheet.cell_value(row, columnno3))  # extract from Altitude col
            '''
            wb = load_workbook(f)
            for sheet1 in wb:
                print("\n")
            for row in sheet1.iter_rows(
                    'A{}:A{}'.format(cell_row1, cell_row2)):  # here A defines the column you want to search
                for col in sheet1.iter_cols(min_col=1,
                                            max_col=1):  # min column and maaximum column diffrence define the repetion of cell value
                    for cell in row:
                        # Getting all cell values from all rows and first column
                        cv = cell.value
                        column0.append(cell.value)
            '''
            book = xlrd.open_workbook(f)
            sheet = book.sheet_by_index(0)
            for row in range(cell_row11, cell_row2):
                column0.append(str(datetime.timedelta(sheet.cell_value(row, columnno0))))



        # frame Extraction

        length = len(column0)
        intial = '15:17:39'
        s2 = str(column0[length - 1])
        s1 = str(column0[0])

        # logic for time diffence and conversion into milisecond

        FMT = '%H:%M:%S'
        tdelta = datetime.datetime.strptime(s1, FMT) - datetime.datetime.strptime(intial, FMT)
        tdelta1 = datetime.datetime.strptime(s2, FMT) - datetime.datetime.strptime(intial, FMT)

        s = str(tdelta)
        r = str(tdelta1)
        hours, minutes, seconds = (["0", "0"] + s.split(":"))[-3:]
        hours = int(hours)
        minutes = int(minutes)
        seconds = float(seconds)
        miliseconds = int(3600000 * hours + 60000 * minutes + 1000 * seconds)

        hours1, minutes1, seconds1 = (["0", "0"] + r.split(":"))[-3:]
        hours1 = int(hours1)
        minutes1 = int(minutes1)
        seconds1 = float(seconds1)
        miliseconds1 = int(3600000 * hours1 + 60000 * minutes1 + 1000 * seconds1)

        # extract frame from video

        video = self.videodata.get()
        folder = str(self.outputfoldername.get())
        cap = cv2.VideoCapture(video)
        success, image = cap.read()
        frameRate = cap.get(5)  # frame rate

        if (str(self.numberofframe.get()) != "All"):

            if (int(self.numberofframe.get()) >= int(frameRate)):
                showerror("Error", "Frame you demand is out of Range")
                self.cmb2.focus_set()
                return

        c = 0
        position = 0
        pos_increament=1
        global counter
        counter=0
        if (self.numberofframe.get() == '1'):
            while (cap.isOpened()):

                frameId = cap.get(1)  # current frame number
                getvalue = cap.get(0)
                ret, frame = cap.read()

                if (ret != True):
                    break
                if (frameId % math.floor(frameRate) == 0):
                    if getvalue >= miliseconds and getvalue <= miliseconds1:
                        filename = folder + "/image_" + str(int(frameId)) + ".jpg"

                        self.label.config(text="Extracting Frame -"+str(int(frameId)),fg="red")
                        self.label.update_idletasks()

                        columnimagename.append("/image_" + str(int(frameId)) + ".jpg")
                        c = c + 1
                        cv2.imwrite(filename, frame)
                        self.set_gps_location(filename, column2[position], column3[position],column4[position])
                        position = position + 1
                        counter=1
            cap.release()

        elif(self.numberofframe.get() == "All"):

            count = 0
            while success:
                success, image = cap.read()

                getvalue = cap.get(0)


                if getvalue >= miliseconds and getvalue <= miliseconds1:
                    cv2.imwrite(os.path.join(folder, "frame{:d}.jpg".format(count)), image)

                    self.label.config(text="Extracting Frame -"+str(count),fg="red")
                    self.label.update_idletasks()


                    filename = folder + "/frame" + str(count) + ".jpg"
                    self.set_gps_location(filename,column2[position],column3[position],column4[position])
                    if pos_increament == int(frameRate):
                        pos_increament = 1
                        position = position + 1
                    else:
                        pos_increament = pos_increament + 1

                    counter=1


                count += 1



            cap.release()
        else:
            totalmilisecond = 1000
            rate = int(self.numberofframe.get())
            param = math.floor(totalmilisecond / rate)
            count = 0
            success, image = cap.read()
            success = True
            while success:
                cap.set(cv2.CAP_PROP_POS_MSEC,
                        (count * param))  # set the position of frames of video acording to milisecond
                success, image = cap.read()
                getvalue = cap.get(0)

                if getvalue >= miliseconds and getvalue <= miliseconds1:



                    self.label.config(text="Extracting frame"+str(count),fg="red")
                    self.label.update_idletasks()


                    cv2.imwrite(os.path.join(folder, "frame{:d}.jpg".format(count)), image)
                    filename = folder + "/frame" + str(count) + ".jpg"
                    self.set_gps_location(filename, column2[position], column3[position],column4[position])
                    if pos_increament == int(self.numberofframe.get()):

                        pos_increament = 1
                        position = position + 1
                    else:
                        pos_increament = pos_increament + 1
                    counter=1
                count = count + 1

            cap.release()



        if counter ==1:
           self.label.config(text="########## PROCESS SUCESSFULLY COMPLETED #########",fg="red",font=('verdana', 10, 'bold'))
           self.label.update_idletasks()
        else:
            self.label.config(text="Try Again",fg="yellow" ,font=('verdana', 10, 'bold'))
            self.label.update_idletasks()




if __name__ == '__main__':
        root = Tk()
        root.resizable(0, 0)
        obj = REG(root)
        root.mainloop()


